import csv
import cPickle as pickle
from openpyxl import Workbook
from openpyxl import load_workbook

def to_ascii(s):
    if isinstance(s, (long, int, float)):
        return s
    try:
        return s.encode("ascii")
    except:
        chars = [c for c in s]
        for i, c in enumerate(chars):
            try:
                c.encode("ascii")
            except:
                chars[i] = " "
        return ''.join(chars).strip()

def read_csv(path, delim=',', stripSpace=True):
    with open(path, "rb") as f:
        reader = csv.reader(f, delimiter=delim)
        if stripSpace:
            return [[to_ascii(e).strip() for e in line] for line in reader]
        else:
            return [[to_ascii(e) for e in line] for line in reader]

def write_csv(path, data):
    with open(path, "wb") as f:
        writer = csv.writer(f)
        for row in data:
            writer.writerow(row)

def read_ws(path, sheet_name):
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet_name]
    def format_value(v):
        if v is None:
            return None
        else:
            return to_ascii(v)

    return [[format_value(cell.value) for cell in row] for row in ws.rows]

def read_wb(path):
    wb = load_workbook(path, read_only=True)
    data = dict()
    for s in wb.get_sheet_names():
        ws = wb[s]
        def format_value(v):
            if v is None:
                return None
            else:
                return to_ascii(v)
        data[s] = [[format_value(cell.value) for cell in row] for row in ws.rows]

    return data

def write_pickle(path, data):
    out_s = open(path, 'wb')
    for obj in data:
        pickle.dump(obj, out_s)
    out_s.close()

def read_pickle(path):
    in_s = open(path, 'rb')
    result = []
    while True:
        try:
            obj = pickle.load(in_s)
            result.append(obj)
        except EOFError:
            break
    return result
    


